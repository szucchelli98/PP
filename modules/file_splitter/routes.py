import io
import os
import zipfile

import openpyxl
from flask import Blueprint, Response, render_template, request

splitter_bp = Blueprint("splitter", __name__, template_folder="templates")

MAX_ROWS = 15_000


@splitter_bp.route("/")
def index():
    return render_template("file_splitter/index.html")


@splitter_bp.route("/split", methods=["POST"])
def split():
    try:
        file = request.files.get("file")
        if not file or file.filename == "":
            return _json_error("No file uploaded", 400)

        ext = os.path.splitext(file.filename)[1].lower()
        if ext not in (".xlsx", ".xls"):
            return _json_error("Only .xlsx / .xls files are supported", 400)

        # ── Read source workbook ──────────────────────────────────────────
        file_bytes = file.read()
        wb_src = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws_src = wb_src.active
        all_rows = [list(row) for row in ws_src.iter_rows(values_only=True)]
        wb_src.close()

        if not all_rows:
            return _json_error("File is empty", 400)

        header = all_rows[0]
        data   = all_rows[1:]
        chunks = [data[i: i + MAX_ROWS] for i in range(0, max(len(data), 1), MAX_ROWS)]
        base   = os.path.splitext(file.filename)[0]

        # ── Build ZIP with one .xlsx per chunk ───────────────────────────
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for i, chunk in enumerate(chunks, 1):
                wb_out = openpyxl.Workbook()
                ws_out = wb_out.active
                ws_out.append(header)
                for row in chunk:
                    ws_out.append(row)
                part_buf = io.BytesIO()
                wb_out.save(part_buf)
                zf.writestr(f"{base}_part{i:02d}.xlsx", part_buf.getvalue())

        zip_bytes = buf.getvalue()
        return Response(
            zip_bytes,
            status=200,
            mimetype="application/zip",
            headers={
                "Content-Disposition": f'attachment; filename="{base}_split.zip"',
                "X-Parts-Count": str(len(chunks)),
                "Content-Length": str(len(zip_bytes)),
            },
        )

    except Exception as exc:
        return _json_error(f"Unexpected error: {exc}", 500)


def _json_error(msg: str, status: int) -> Response:
    safe = msg.replace('"', "'")
    return Response(f'{{"error":"{safe}"}}', status=status, mimetype="application/json")
